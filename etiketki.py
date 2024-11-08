import time
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from collections import defaultdict
from openpyxl import load_workbook
import os


# Настройки
EXCEL_INPUT_FILE = 'Артикли.xlsx'        # Входной Excel файл
EXCEL_OUTPUT_FILE = 'Результаты.xlsx'    # Выходной Excel файл
SEARCH_URL = 'https://kirpich.ru/shop'

# Чтение данных из Excel
def read_articles(file_path):
    df = pd.read_excel(file_path)
    # Предполагаем, что нужные артикулы находятся в первом столбце
    articles = df.iloc[:, 0].dropna().astype(str).tolist()
    return articles

# Запись данных в Excel
def write_results(data, file_path):
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False)

# Инициализация веб-драйвера
def init_driver():
    options = webdriver.ChromeOptions()
    #options.add_argument('--headless')  # Запуск без интерфейса браузера
    options.add_argument('--disable-gpu')
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    return driver

# Основная функция
def main():
    articles = read_articles(EXCEL_INPUT_FILE)
    driver = init_driver()
    results = []
    specs = {}
    all_specs = set()  # Для сбора всех возможных спецификаций

    for idx, article in enumerate(articles, 1):
        try:
            print(f"Обработка {idx}/{len(articles)}: {article}")
            driver.get(SEARCH_URL)
            time.sleep(2)  # Ждем загрузки страницы

            # Найти поле поиска
            search_box = driver.find_element(By.CSS_SELECTOR, 'input[name="q"]')
            search_box.clear()
            search_box.send_keys(article)
            search_box.send_keys(Keys.RETURN)  # Нажать Enter
            time.sleep(3)  # Ждем результатов поиска

            # Выбрать первый продукт в результатах
            first_product = driver.find_element(By.CSS_SELECTOR, 'a.card__slider')
            first_product.click()
            time.sleep(3)  # Ждем загрузки страницы продукта

            # Находим все элементы <li> внутри <ul class="specs-list">
            li_elements = driver.find_elements(By.CSS_SELECTOR, 'ul.chars__list.chars__list--dotted li')
            
            for li in li_elements:
                try:
                    # Извлекаем все <span> внутри текущего <li>
                    spans = li.find_elements(By.TAG_NAME, 'span')
                    
                    if len(spans) >= 2:
                        # Первый <span> с классом "product_content_info" — это ключ
                        key = spans[0].text.strip()
                        # Второй <span> — это значение
                        value = spans[1].text.strip()
                        
                        # Записываем в словарь specs
                        specs[key] = value
                        # Добавляем ключ в множество all_specs
                        all_specs.add(key)
                    else:
                        # Если по какой-то причине количество <span> меньше 2, обрабатываем иначе
                        text = li.text.strip()
                        specs[text] = None
                        all_specs.add(text)
                except Exception as e:
                    print(f"Ошибка при обработке элемента: {e}")

        except Exception as e:
            print(f"Ошибка при обработке артикла {article}: {e}")
            result = {'Артикул': article, 'Ошибка': str(e)}
            results.append(result)

    # Закрыть драйвер
    driver.quit()

    # Создать DataFrame с учетом всех спецификаций
    df = pd.DataFrame([specs])

    # Переупорядочить столбцы: сначала 'Артикул', затем спецификации в алфавитном порядке или нужном порядке
    #cols = ['Артикул'] + list(all_specs)
    #if 'Ошибка' in df.columns:
    #    cols.append('Ошибка')
    #df = df.reindex(columns=cols)

    # Записать результаты в Excel
    #write_results(df, EXCEL_OUTPUT_FILE)
    #print(f"Завершено. Результаты сохранены в {EXCEL_OUTPUT_FILE}")

    # Проверяем, существует ли файл
    if os.path.exists(EXCEL_OUTPUT_FILE):
        # Загружаем существующий файл
        book = load_workbook(EXCEL_OUTPUT_FILE)
        writer = pd.ExcelWriter(EXCEL_OUTPUT_FILE, engine='openpyxl') 
        writer.book = book

        # Если лист с нужным именем уже существует, выбираем его
        if 'Sheet1' in book.sheetnames:
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            # Определяем, с какой строки начать запись
            startrow = writer.sheets['Sheet1'].max_row

            # Записываем новые данные без перезаписи заголовков
            df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=False)
        else:
            # Если листа нет, создаем его с заголовками
            df.to_excel(writer, sheet_name='Sheet1', index=False)

        writer.save()
    else:
        # Если файла не существует, создаем его и записываем данные с заголовками
        with pd.ExcelWriter(EXCEL_OUTPUT_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)

    print("Данные успешно добавлены в Excel файл.")

if __name__ == "__main__":
    main()
